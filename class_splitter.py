# -*- coding: utf-8 -*-
"""Balanced random class splitter for Excel student lists."""

from __future__ import annotations

import argparse
import datetime as dt
import gc
import math
import random
import re
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:  # pragma: no cover - this is a friendly CLI failure path.
    raise SystemExit(
        "缺少 openpyxl。请先安装：python -m pip install openpyxl"
    ) from exc


REQUIRED_COLUMNS = ["姓名", "性别"]
OPTIONAL_COLUMNS = ["是否定向", "指定班级", "户籍细类", "户籍大类", "备注"]
OFFICIAL_TEMPLATE_COLUMNS = [
    "总人数序号",
    "学校序号",
    "姓名",
    "证件类型",
    "证件号码",
    "儿童人员类型",
    "报名情况",
    "分配验证点",
    "性别",
    "出生日期",
    "国籍",
    "民族",
    "是否持上海市居住证",
    "居住证类型",
    "是否就读幼儿园",
    "就读幼儿园名称",
    "户籍地址(省)",
    "户籍地址(区)",
    "户籍地址(街道)",
    "户籍地址(居委)",
    "户籍地详址",
    "与户主关系",
    "户主姓名",
    "户主证件号码",
    "户口性质",
    "户籍登记日",
    "居住地址(区)",
    "居住地址(街镇)",
    "居住地址(居委村)",
    "居住地详址",
    "儿童人户分离登记情况",
    "住房与地址关系",
    "住房性质",
    "产证类型/其他类型",
    "产权人姓名",
    "产权编号/合同编号",
    "登记日期",
    "与购房人或承租人关系",
    "监护人姓名",
    "监护人性别",
    "与监护人关系",
    "监护人手机号码",
    "监护人证件类型",
    "监护人证件号码",
    "监护人学历",
    "监护人工作单位",
    "监护人户籍所在地(省)",
    "监护人户籍所在地(区)",
    "第二监护人姓名",
    "第二监护人性别",
    "与第二监护人关系",
    "第二监护人手机号码",
    "第二监护人证件类型",
    "第二监护人证件号码",
    "第二监护人学历",
    "第二监护人工作单位",
    "第二监护人户籍所在地(省)",
    "第二监护人户籍所在地(区)",
    "监护人是否持有消防员证/消防干部证",
    "港澳台侨外",
    "是否领取残疾证",
    "残疾类别",
    "残疾证编号",
    "残疾证发证日期",
    "残疾证发证机关",
    "是否领取阳光宝宝卡",
    "阳光宝宝卡残疾类别",
    "阳光宝宝卡发证日期",
    "备注1",
    "备注2",
    "备注3",
    "备注4",
    "登记点名称",
    "登记点工作人员姓名",
    "登记时间",
    "登记号",
    "对口学校",
    "录取学校",
]
TOOL_EXTRA_TEMPLATE_COLUMNS = ["是否定向", "指定班级", "户籍大类", "户籍细类"]
TEMPLATE_COLUMNS = OFFICIAL_TEMPLATE_COLUMNS + TOOL_EXTRA_TEMPLATE_COLUMNS

HUKOU_DETAILS = ["本市人户一致", "本市人户分离", "外省市满积分", "外省市不满积分"]
HUKOU_MAIN_BY_DETAIL = {
    "本市人户一致": "上海市户口",
    "本市人户分离": "上海市户口",
    "外省市满积分": "外省市户口",
    "外省市不满积分": "外省市户口",
}
HUKOU_MAIN_VALUES = ["上海市户口", "外省市户口"]
GENDER_VALUES = ["男", "女"]
YES_NO_VALUES = ["否", "是"]
DEFAULT_BALANCE_OPTIONS = "ABC"
BALANCE_OPTIONS = {
    "A": "性别",
    "B": "户籍大类",
    "C": "户籍细类",
    "D": "儿童人员类型",
}

HEADER_ALIASES = {
    "学号": ["学号", "学生编号", "编号", "学籍号", "学生号", "总人数序号", "学校序号", "登记号"],
    "姓名": ["姓名", "学生姓名", "名字"],
    "性别": ["性别", "男女"],
    "户籍大类": ["户籍大类", "户口大类", "户籍", "户口", "户籍类型"],
    "户籍细类": ["户籍细类", "户口细类", "户籍信息", "户口信息", "户籍类别", "户口类别"],
    "是否定向": ["是否定向", "定向", "是否指定", "是否指定班级"],
    "指定班级": ["指定班级", "定向班级", "目标班级", "安排班级"],
    "备注": ["备注", "说明"],
    "儿童人员类型": ["儿童人员类型"],
    "报名情况": ["报名情况"],
    "分配验证点": ["分配验证点"],
    "户籍地址省": ["户籍地址(省)", "户籍地址（省）", "户籍省", "户籍所在地(省)", "户籍所在地（省）"],
    "户籍地址区": ["户籍地址(区)", "户籍地址（区）", "户籍区", "户籍所在地(区)", "户籍所在地（区）"],
    "户籍地址街道": ["户籍地址(街道)", "户籍地址（街道）", "户籍街道", "户籍地址(街镇)", "户籍地址（街镇）"],
    "户籍地址居委": ["户籍地址(居委)", "户籍地址（居委）", "户籍居委", "户籍地址(居委村)", "户籍地址（居委村）"],
    "户籍地详址": ["户籍地详址", "户籍详细地址", "户籍地址"],
    "户口性质": ["户口性质"],
    "居住地址区": ["居住地址(区)", "居住地址（区）", "居住区"],
    "居住地址街镇": ["居住地址(街镇)", "居住地址（街镇）", "居住街镇", "居住地址(街道)", "居住地址（街道）"],
    "居住地址居委村": ["居住地址(居委村)", "居住地址（居委村）", "居住居委村", "居住地址(居委)", "居住地址（居委）"],
    "居住地详址": ["居住地详址", "居住详细地址", "居住地址"],
    "儿童人户分离登记情况": ["儿童人户分离登记情况", "人户分离登记情况", "人户分离"],
    "是否持上海市居住证": ["是否持上海市居住证"],
    "居住证类型": ["居住证类型"],
    "备注1": ["备注1"],
    "备注2": ["备注2"],
    "备注3": ["备注3"],
    "备注4": ["备注4"],
}

RESULT_SHEET_NAME = "分班结果"
EXTRA_RESULT_COLUMNS = ["最终班级", "班内序号", "推断户籍大类", "推断户籍细类", "分层类别", "分班方式", "未参与原因", "原始行号"]


@dataclass
class Issue:
    level: str
    row_number: int | str
    field: str
    message: str


@dataclass
class Student:
    row_number: int
    original_values: dict[str, Any]
    canonical_values: dict[str, Any]
    student_id: str
    name: str
    gender: str
    hukou_main: str
    hukou_detail: str
    directed: bool
    target_class: str
    assigned_class: str = ""
    assignment_type: str = "自动"
    participates: bool = True
    skip_reason: str = ""

    @property
    def layer(self) -> str:
        return f"{self.gender}_{self.hukou_detail}"


@dataclass
class SplitResult:
    output_path: Path
    student_count: int
    class_names: list[str]
    seed: int
    balance_options: str
    issues: list[Issue]


class ClassSplitError(Exception):
    def __init__(self, message: str, report_path: Path | None = None):
        super().__init__(message)
        self.report_path = report_path


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: Any) -> str:
    text = clean(value)
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def split_csv_like(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in re.split(r"[,，;；、\n]+", value) if item.strip()]


def canonical_header(value: Any) -> str:
    key = compact(value)
    if not key:
        return ""
    for canonical, aliases in HEADER_ALIASES.items():
        if key in {compact(alias) for alias in aliases}:
            return canonical
    return clean(value)


def normalize_gender(value: Any) -> str:
    key = compact(value).lower()
    if "男" in key:
        return "男"
    if "女" in key:
        return "女"
    if key in {"m", "male", "1"}:
        return "男"
    if key in {"f", "female", "0", "2"}:
        return "女"
    return clean(value)


def normalize_hukou_detail(value: Any) -> str:
    key = compact(value)
    aliases = {
        "本市人户一致": "本市人户一致",
        "人户一致": "本市人户一致",
        "上海人户一致": "本市人户一致",
        "本市人户分离": "本市人户分离",
        "人户分离": "本市人户分离",
        "上海人户分离": "本市人户分离",
        "外省市满积分": "外省市满积分",
        "外省满积分": "外省市满积分",
        "满积分": "外省市满积分",
        "积分达标": "外省市满积分",
        "外省市不满积分": "外省市不满积分",
        "外省不满积分": "外省市不满积分",
        "不满积分": "外省市不满积分",
        "积分不达标": "外省市不满积分",
    }
    return aliases.get(key, clean(value))


def normalize_hukou_main(value: Any, detail: str) -> str:
    key = compact(value)
    if detail in HUKOU_MAIN_BY_DETAIL:
        inferred = HUKOU_MAIN_BY_DETAIL[detail]
        if not key:
            return inferred
    aliases = {
        "上海市户口": "上海市户口",
        "上海户口": "上海市户口",
        "本市户口": "上海市户口",
        "本市": "上海市户口",
        "外省市户口": "外省市户口",
        "外省户口": "外省市户口",
        "外地户口": "外省市户口",
        "外省市": "外省市户口",
        "外地": "外省市户口",
    }
    return aliases.get(key, clean(value))


def is_directed(value: Any, target_class: str) -> bool:
    key = compact(value).lower()
    if target_class:
        return True
    return key in {"是", "y", "yes", "true", "1", "√", "对", "指定", "定向"}


def get_value(values: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean(values.get(key))
        if value:
            return value
    return ""


def joined_values(values: dict[str, Any], keys: list[str]) -> str:
    return " ".join(clean(values.get(key)) for key in keys if clean(values.get(key)))


def has_any(text: str, keywords: list[str]) -> bool:
    compact_text = compact(text)
    return any(keyword in compact_text for keyword in keywords)


def is_shanghai_province(value: str) -> bool:
    key = compact(value)
    return key in {"上海", "上海市", "沪", "31", "310000"} or "上海" in key


def same_place(values: dict[str, Any], left_keys: list[str], right_keys: list[str]) -> bool | None:
    pairs = [(compact(values.get(left)), compact(values.get(right))) for left, right in zip(left_keys, right_keys)]
    pairs = [(left, right) for left, right in pairs if left and right]
    if not pairs:
        return None
    return all(left == right for left, right in pairs)


def infer_shanghai_detail(values: dict[str, Any], issues: list[Issue], row_number: int) -> str:
    separation_text = joined_values(values, ["儿童人户分离登记情况", "分配验证点", "儿童人员类型", "报名情况"])
    if has_any(separation_text, ["人户分离", "已登记", "分离", "不一致"]):
        return "本市人户分离"
    if has_any(separation_text, ["人户一致", "未登记", "未办理", "无需登记", "否", "无"]):
        return "本市人户一致"

    same_district = same_place(values, ["户籍地址区"], ["居住地址区"])
    same_street = same_place(values, ["户籍地址街道"], ["居住地址街镇"])
    same_committee = same_place(values, ["户籍地址居委"], ["居住地址居委村"])
    if same_district is False or same_street is False or same_committee is False:
        return "本市人户分离"
    if all(item is True for item in [same_district, same_street, same_committee] if item is not None):
        return "本市人户一致"

    issues.append(
        Issue(
            "警告",
            row_number,
            "户籍细类",
            "未明确看到人户一致/分离信息，已默认按“本市人户一致”处理；可新增或填写“户籍细类”列覆盖。",
        )
    )
    return "本市人户一致"


def infer_non_shanghai_detail(values: dict[str, Any], issues: list[Issue], row_number: int, required: bool) -> str:
    score_text = joined_values(
        values,
        [
            "户籍细类",
            "儿童人员类型",
            "居住证类型",
            "是否持上海市居住证",
            "报名情况",
            "分配验证点",
            "备注",
            "备注1",
            "备注2",
            "备注3",
            "备注4",
        ],
    )
    if has_any(score_text, ["不满积分", "未满积分", "积分不达标", "未达标", "未达到", "不达标", "未积分"]):
        return "外省市不满积分"
    if has_any(score_text, ["外省市满积分", "满积分", "积分达标", "达到标准分值", "标准分值", "积分达到"]):
        return "外省市满积分"
    issues.append(
        Issue(
            "错误" if required else "警告",
            row_number,
            "户籍细类",
            "外省市户口未能判断满积分/不满积分；如需按 C=户籍细类 分班，请在“儿童人员类型”“居住证类型”中保留相关关键词，或新增“户籍细类”列填写。",
        )
    )
    return ""


def infer_hukou(
    values: dict[str, Any],
    issues: list[Issue],
    row_number: int,
    require_main: bool,
    require_detail: bool,
) -> tuple[str, str]:
    explicit_detail = normalize_hukou_detail(values.get("户籍细类"))
    if explicit_detail in HUKOU_DETAILS:
        return HUKOU_MAIN_BY_DETAIL[explicit_detail], explicit_detail

    explicit_main = normalize_hukou_main(values.get("户籍大类"), "")
    province = get_value(values, "户籍地址省")
    account_text = joined_values(values, ["户口性质", "儿童人员类型", "报名情况", "分配验证点"])

    if explicit_main == "上海市户口" or is_shanghai_province(province) or has_any(account_text, ["本市户籍", "上海户籍", "上海市户籍"]):
        return "上海市户口", infer_shanghai_detail(values, issues, row_number)
    if explicit_main == "外省市户口" or province or has_any(account_text, ["外省市", "外省", "外地", "非本市", "积分"]):
        return "外省市户口", infer_non_shanghai_detail(values, issues, row_number, require_detail)

    issues.append(
        Issue(
            "错误" if require_main else "警告",
            row_number,
            "户籍地址(省)",
            "未能判断上海市户口或外省市户口；如需按 B/C 分班，请检查“户籍地址(省)”或新增“户籍大类/户籍细类”列。",
        )
    )
    return "", ""


def parse_balance_options(value: str | None) -> str:
    raw = compact(value or DEFAULT_BALANCE_OPTIONS).upper()
    options = "".join(char for char in raw if char.isalpha())
    if not options:
        options = DEFAULT_BALANCE_OPTIONS
    invalid = [char for char in options if char not in BALANCE_OPTIONS]
    if invalid:
        raise ClassSplitError(
            f"分班依据只能输入 {''.join(BALANCE_OPTIONS)} 中的字母，例如 A、AB、ABC、ABCD。"
        )
    deduped = ""
    for char in options:
        if char not in deduped:
            deduped += char
    return deduped


def balance_option_text(options: str) -> str:
    return "，".join(f"{char}={BALANCE_OPTIONS[char]}" for char in options)


def balance_option_names(options: str) -> list[str]:
    return [BALANCE_OPTIONS[option] for option in options if option in BALANCE_OPTIONS]


def human_join(items: list[str]) -> str:
    if not items:
        return "未指定项目"
    if len(items) == 1:
        return items[0]
    return "、".join(items)


def plain_language_summary(
    balance_options: str,
    students: list[Student],
    participating_students: list[Student],
    skipped_students: list[Student],
    class_names: list[str],
    capacity: int | None,
    notices: list[str] | None,
) -> list[str]:
    basis = human_join(balance_option_names(balance_options))
    directed_count = sum(1 for student in participating_students if student.directed)
    lines = [
        f"本次分班一共读取 {len(students)} 行学生记录，其中 {len(participating_students)} 人参与自动分班，{len(skipped_students)} 人因资料缺失或格式问题未参与分班。",
        f"本次主要根据 {basis} 尽量平均分到 {len(class_names)} 个班：{'，'.join(class_names)}。",
        "做法是先把需要定向安排的学生放入指定班级，再把其他学生按所选依据分成小组，随机打乱后尽量平均分配到各班。",
    ]
    if directed_count:
        lines.append(f"其中有 {directed_count} 名学生属于定向安排，系统会优先尊重指定班级。")
    if capacity is not None:
        lines.append(f"每班人数上限设置为 {capacity} 人；如果原班级容量不够，系统会自动增加班级。")
    if notices:
        lines.extend(notices)
    if skipped_students:
        lines.append("未参与分班的学生已放在总表末尾和“未参与分班”分块中，可补齐信息后重新分班。")
    lines.append("随机种子记录在运行信息中；名单和设置不变时，使用同一个随机种子可以复现同一套分班结果。")
    return lines


def balance_value(student: Student, option: str) -> str:
    if option == "A":
        return student.gender or "未填性别"
    if option == "B":
        return student.hukou_main or "未判断户籍大类"
    if option == "C":
        return student.hukou_detail or "未判断户籍细类"
    if option == "D":
        return clean(student.canonical_values.get("儿童人员类型")) or "未填儿童人员类型"
    return ""


def balance_layer(student: Student, options: str) -> str:
    return " | ".join(f"{BALANCE_OPTIONS[option]}:{balance_value(student, option)}" for option in options)


def int_or_none(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def find_header(ws, requested_sheet: str | None = None) -> tuple[int, list[int], list[str], list[str]]:
    max_row = min(ws.max_row or 20, 20)
    required = set(REQUIRED_COLUMNS)
    for row, raw_headers_tuple in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        raw_headers = list(raw_headers_tuple)
        canonical_headers = [canonical_header(value) for value in raw_headers]
        if required.issubset(set(canonical_headers)):
            used_cols: list[int] = []
            used_raw: list[str] = []
            used_canonical: list[str] = []
            for idx, (raw, canonical) in enumerate(zip(raw_headers, canonical_headers), start=1):
                if clean(raw):
                    used_cols.append(idx)
                    used_raw.append(clean(raw))
                    used_canonical.append(canonical)
            return row, used_cols, used_raw, used_canonical
    sheet_hint = f"工作表“{requested_sheet}”" if requested_sheet else "名单工作表"
    raise ClassSplitError(
        f"{sheet_hint}没有找到表头。第一行或前20行里需要包含：{', '.join(REQUIRED_COLUMNS)}"
    )


def choose_data_sheet(wb, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ClassSplitError(f"找不到工作表：{sheet_name}")
        return wb[sheet_name]
    if "学生名单" in wb.sheetnames:
        return wb["学生名单"]
    return wb.active


def read_config(wb) -> dict[str, str]:
    if "配置" not in wb.sheetnames:
        return {}
    ws = wb["配置"]
    config: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row=row, column=1).value)
        value = clean(ws.cell(row=row, column=2).value)
        if key:
            config[key] = value
    return config


def resolve_class_names(
    config: dict[str, str],
    class_count: int | None,
    class_names: list[str] | None,
) -> list[str]:
    config_names = split_csv_like(config.get("班级名称", ""))
    config_count = int_or_none(config.get("班级数量", ""))

    names = class_names or config_names
    if names:
        if len(set(names)) != len(names):
            raise ClassSplitError("班级名称不能重复。")
        if class_count and len(names) != class_count:
            raise ClassSplitError("班级数量和班级名称个数不一致。")
        return names

    count = class_count or config_count or 4
    if count < 1:
        raise ClassSplitError("班级数量必须大于0。")
    return [f"{idx}班" for idx in range(1, count + 1)]


def next_class_name(class_names: list[str]) -> str:
    used = set(class_names)
    numbers = []
    for name in class_names:
        match = re.fullmatch(r"(\d+)班", clean(name))
        if match:
            numbers.append(int(match.group(1)))
    candidate_number = (max(numbers) + 1) if numbers else (len(class_names) + 1)
    candidate = f"{candidate_number}班"
    while candidate in used:
        candidate_number += 1
        candidate = f"{candidate_number}班"
    return candidate


def expand_classes_for_capacity(
    class_names: list[str],
    capacity: int | None,
    student_count: int,
) -> tuple[list[str], list[str]]:
    if capacity is None or capacity <= 0:
        return class_names, []
    expanded = list(class_names)
    added = []
    while capacity * len(expanded) < student_count:
        new_name = next_class_name(expanded)
        expanded.append(new_name)
        added.append(new_name)
    if not added:
        return expanded, []
    message = (
        f"因每班上限为{capacity}人，原{len(class_names)}个班最多容纳{capacity * len(class_names)}人，"
        f"参与分班学生为{student_count}人，已自动新增班级：{'，'.join(added)}。"
    )
    return expanded, [message]


def read_students(
    input_path: Path,
    sheet_name: str | None,
    class_names: list[str],
    balance_options: str,
) -> tuple[list[Student], list[str], list[Issue]]:
    wb = load_workbook(input_path, read_only=True, data_only=False)
    try:
        ws = choose_data_sheet(wb, sheet_name)
        header_row, used_cols, raw_headers, canonical_headers = find_header(ws, sheet_name)
    except Exception:
        wb.close()
        raise

    students: list[Student] = []
    issues: list[Issue] = []
    class_set = set(class_names)
    require_main = "B" in balance_options or "C" in balance_options
    require_detail = "C" in balance_options

    for row, row_values_tuple in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        row_values = list(row_values_tuple)
        original_values: dict[str, Any] = {}
        canonical_values: dict[str, Any] = {}
        row_issues: list[Issue] = []
        for col, raw_header, canonical in zip(used_cols, raw_headers, canonical_headers):
            value = row_values[col - 1] if col - 1 < len(row_values) else None
            original_values[raw_header] = value
            if canonical not in canonical_values or not clean(canonical_values.get(canonical)):
                canonical_values[canonical] = value

        if not any(clean(value) for value in original_values.values()):
            continue

        student_id = clean(canonical_values.get("学号")) or str(row - header_row)
        name = clean(canonical_values.get("姓名"))
        gender = normalize_gender(canonical_values.get("性别"))
        main, detail = infer_hukou(canonical_values, row_issues, row, require_main, require_detail)
        target = clean(canonical_values.get("指定班级"))
        directed = is_directed(canonical_values.get("是否定向"), target)

        if not name:
            row_issues.append(Issue("错误", row, "姓名", "姓名不能为空。"))
        if gender not in GENDER_VALUES:
            row_issues.append(Issue("错误", row, "性别", f"性别应为：{', '.join(GENDER_VALUES)}。"))
        if detail and detail not in HUKOU_DETAILS:
            row_issues.append(Issue("错误", row, "户籍细类", f"户籍细类应为：{', '.join(HUKOU_DETAILS)}。"))
        if main and main not in HUKOU_MAIN_VALUES:
            row_issues.append(Issue("错误", row, "户籍大类", f"户籍大类应为：{', '.join(HUKOU_MAIN_VALUES)}。"))
        if directed and not target:
            row_issues.append(Issue("错误", row, "指定班级", "是否定向为“是”时，必须填写指定班级。"))
        if target and target not in class_set:
            row_issues.append(Issue("错误", row, "指定班级", f"指定班级“{target}”不在班级列表中。"))

        error_messages = [f"{issue.field}: {issue.message}" for issue in row_issues if issue.level == "错误"]
        participates = not error_messages
        skip_reason = "；".join(error_messages)
        if not participates:
            for issue in row_issues:
                if issue.level == "错误":
                    issue.level = "跳过"
            if not skip_reason:
                skip_reason = "关键信息缺失，未参与分班。"
        issues.extend(row_issues)

        students.append(
            Student(
                row_number=row,
                original_values=original_values,
                canonical_values=canonical_values,
                student_id=student_id,
                name=name,
                gender=gender,
                hukou_main=main,
                hukou_detail=detail,
                directed=directed,
                target_class=target,
                assigned_class="" if participates else "未参与分班",
                assignment_type="自动" if participates else "未参与",
                participates=participates,
                skip_reason=skip_reason,
            )
        )

    if not students:
        issues.append(Issue("错误", "-", "学生名单", "没有读取到学生数据。"))

    wb.close()
    gc.collect()
    return students, raw_headers, issues


def target_class_sizes(total: int, class_names: list[str]) -> dict[str, int]:
    base = total // len(class_names)
    extra = total % len(class_names)
    return {name: base + (1 if idx < extra else 0) for idx, name in enumerate(class_names)}


def assign_students(
    students: list[Student],
    class_names: list[str],
    seed: int,
    capacity: int | None,
    balance_options: str,
    issues: list[Issue],
) -> None:
    rng = random.Random(seed)
    class_counts: Counter[str] = Counter()
    layer_counts: dict[str, Counter[str]] = {name: Counter() for name in class_names}
    targets = target_class_sizes(len(students), class_names)

    for student in students:
        if not student.directed:
            continue
        student.assigned_class = student.target_class
        student.assignment_type = "定向"
        class_counts[student.assigned_class] += 1
        layer_counts[student.assigned_class][balance_layer(student, balance_options)] += 1

    if capacity is not None:
        for class_name, count in class_counts.items():
            if count > capacity:
                issues.append(
                    Issue("错误", "-", "每班人数上限", f"{class_name}定向人数为{count}，已超过上限{capacity}。")
                )

    groups: dict[str, list[Student]] = defaultdict(list)
    for student in students:
        if not student.directed:
            groups[balance_layer(student, balance_options)].append(student)

    ordered_groups = sorted(groups.items(), key=lambda item: (-len(item[1]), item[0]))
    for layer, members in ordered_groups:
        rng.shuffle(members)
        for student in members:
            candidates = []
            for class_name in class_names:
                if capacity is not None and class_counts[class_name] >= capacity:
                    continue
                target = max(targets[class_name], 1)
                over_target = max(0, class_counts[class_name] - targets[class_name])
                fill_ratio = class_counts[class_name] / target
                candidates.append(
                    (
                        layer_counts[class_name][layer],
                        over_target,
                        fill_ratio,
                        class_counts[class_name],
                        rng.random(),
                        class_name,
                    )
                )
            if not candidates:
                issues.append(Issue("错误", student.row_number, "班级容量", "所有班级都已达到人数上限，无法继续分配。"))
                continue
            candidates.sort()
            chosen = candidates[0][-1]
            student.assigned_class = chosen
            student.assignment_type = "自动"
            class_counts[chosen] += 1
            layer_counts[chosen][layer] += 1


def has_errors(issues: list[Issue]) -> bool:
    return any(issue.level == "错误" for issue in issues)


def add_header_style(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9E2F3")
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_title(ws, row: int, title: str, max_col: int = 8) -> None:
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24


def apply_table_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    side = Side(style="thin", color="D9E2F3")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def autofit(ws, min_width: int = 10, max_width: int = 28) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            value = clean(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 4))
        ws.column_dimensions[letter].width = width


def freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def load_workbook_for_write(path: Path):
    # Only the generated result sheet is edited. keep_vba preserves macros when the source is xlsm.
    return load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")


def default_output_path(input_path: Path) -> Path:
    suffix = "_重新分班结果" if "分班结果" in input_path.stem else "_分班结果"
    return input_path.with_name(f"{input_path.stem}{suffix}.xlsx")


def timestamped_path(path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def save_workbook_safely(
    wb: Workbook,
    output_path: Path,
    fallback_dir: Path | None = None,
    allow_alternate: bool = True,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError as exc:
        if not allow_alternate:
            raise ClassSplitError(
                f"无法写入原 Excel 文件：{output_path}。请先关闭这个 Excel 文件，再重新开始分班。"
            ) from exc
        first_alt = timestamped_path(output_path)
        try:
            wb.save(first_alt)
            return first_alt
        except PermissionError:
            if fallback_dir is None:
                raise ClassSplitError(
                    f"无法写入输出文件：{output_path}。请关闭已打开的同名 Excel，或换一个保存位置。"
                ) from exc
            fallback_dir.mkdir(parents=True, exist_ok=True)
            fallback_path = fallback_dir / first_alt.name
            try:
                wb.save(fallback_path)
                return fallback_path
            except PermissionError as fallback_exc:
                raise ClassSplitError(
                    f"无法写入输出文件：{output_path}。请关闭已打开的同名 Excel，或换一个保存位置。"
                ) from fallback_exc


def write_issues_sheet(wb: Workbook, issues: list[Issue], title: str = "异常校验") -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(["操作", "原始行号", "学生姓名", "字段", "说明"])
    for issue in issues:
        ws.append([issue.level, issue.row_number, "", issue.field, issue.message])
    add_header_style(ws)
    freeze_and_filter(ws)
    autofit(ws, max_width=60)


def student_name_by_row(students: list[Student]) -> dict[int, str]:
    return {
        student.row_number: student.name or clean(student.original_values.get("姓名"))
        for student in students
    }


def issue_student_name(issue: Issue, names_by_row: dict[int, str]) -> str:
    if isinstance(issue.row_number, int):
        return names_by_row.get(issue.row_number, "")
    row_number = int_or_none(issue.row_number)
    if row_number is None:
        return ""
    return names_by_row.get(row_number, "")


def write_validation_report(
    output_path: Path,
    raw_headers: list[str],
    students: list[Student],
    issues: list[Issue],
    fallback_dir: Path | None = None,
) -> Path:
    output_path = Path(output_path)
    wb = load_workbook_for_write(output_path) if output_path.exists() else Workbook()
    if RESULT_SHEET_NAME in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[RESULT_SHEET_NAME]
    ws = wb.create_sheet(RESULT_SHEET_NAME)
    ws.append(["分班未完成"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["操作", "原始行号", "学生姓名", "字段", "说明"])
    issue_header_row = ws.max_row
    names_by_row = student_name_by_row(students)
    for issue in issues:
        ws.append([issue.level, issue.row_number, issue_student_name(issue, names_by_row), issue.field, issue.message])
    add_header_style(ws, issue_header_row)
    autofit(ws, max_width=80)
    result_path = save_workbook_safely(wb, output_path, fallback_dir, allow_alternate=False)
    wb.close()
    gc.collect()
    return result_path


def safe_sheet_title(value: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]\:\*\?\/\\]", "_", value).strip() or "班级"
    title = title[:31]
    base = title
    idx = 2
    while title in used:
        suffix = f"_{idx}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used.add(title)
    return title


def roster_headers(original_headers: list[str]) -> list[str]:
    headers = ["班级", "班内序号"]
    preferred = ["姓名", "性别", "推断户籍大类", "推断户籍细类", "分层类别", "分班方式", "未参与原因", "原始行号"]
    for header in preferred:
        if header not in headers:
            headers.append(header)
    for header in original_headers:
        if header not in headers:
            headers.append(header)
    return headers


def roster_row(student: Student, headers: list[str], class_index: int, balance_options: str) -> list[Any]:
    row = []
    for header in headers:
        if header == "班级":
            row.append(student.assigned_class)
        elif header == "班内序号":
            row.append(class_index)
        elif header == "推断户籍大类":
            row.append(student.hukou_main)
        elif header == "推断户籍细类":
            row.append(student.hukou_detail)
        elif header == "分层类别":
            row.append(balance_layer(student, balance_options))
        elif header == "分班方式":
            row.append(student.assignment_type)
        elif header == "未参与原因":
            row.append(student.skip_reason)
        elif header == "原始行号":
            row.append(student.row_number)
        elif header == "姓名":
            row.append(student.name)
        elif header == "性别":
            row.append(student.gender)
        else:
            row.append(student.original_values.get(header))
    return row


def write_roster_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    students: list[Student],
    balance_options: str,
    used_titles: set[str],
) -> None:
    ws = wb.create_sheet(safe_sheet_title(title, used_titles))
    ws.append(headers)
    for idx, student in enumerate(students, start=1):
        ws.append(roster_row(student, headers, idx, balance_options))
    add_header_style(ws)
    freeze_and_filter(ws)
    autofit(ws)


def result_headers_for(original_headers: list[str]) -> list[str]:
    headers = ["最终班级"]
    for header in original_headers:
        if header not in headers:
            headers.append(header)
    for column in EXTRA_RESULT_COLUMNS:
        if column not in headers:
            headers.append(column)
    return headers


def result_row_values(
    student: Student,
    headers: list[str],
    class_index: int | str,
    balance_options: str,
) -> list[Any]:
    row_values = []
    for header in headers:
        if header == "最终班级":
            row_values.append(student.assigned_class)
        elif header == "班内序号":
            row_values.append(class_index)
        elif header == "推断户籍大类":
            row_values.append(student.hukou_main)
        elif header == "推断户籍细类":
            row_values.append(student.hukou_detail)
        elif header == "分层类别":
            row_values.append(balance_layer(student, balance_options))
        elif header == "分班方式":
            row_values.append(student.assignment_type)
        elif header == "未参与原因":
            row_values.append(student.skip_reason)
        elif header == "原始行号":
            row_values.append(student.row_number)
        else:
            row_values.append(student.original_values.get(header))
    return row_values


def compact_roster_headers() -> list[str]:
    return ["班级", "班内序号", "姓名", "性别", "户籍大类", "户籍细类", "分班依据", "分班方式", "未参与原因", "原始行号"]


def compact_roster_values(student: Student, class_index: int | str, balance_options: str) -> list[Any]:
    return [
        student.assigned_class,
        class_index,
        student.name,
        student.gender,
        student.hukou_main,
        student.hukou_detail,
        balance_layer(student, balance_options),
        student.assignment_type,
        student.skip_reason,
        student.row_number,
    ]


def pie_chart_size(category_count: int, charts_in_row: int) -> tuple[float, float]:
    category_count = max(1, category_count)
    charts_in_row = max(1, charts_in_row)
    category_extra = min(max(category_count - 2, 0) * 0.25, 0.8)
    row_pressure = min(max(charts_in_row - 1, 0) * 0.18, 0.55)
    width = 5.0 + category_extra - row_pressure
    height = 4.5 + min(max(category_count - 2, 0) * 0.22, 0.9)
    if charts_in_row >= 4:
        height -= 0.15
    return round(min(5.8, max(4.6, width)), 1), round(min(5.4, max(4.3, height)), 1)


def pie_chart_anchor_step(charts_in_row: int) -> int:
    if charts_in_row <= 1:
        return 1
    if charts_in_row == 2:
        return 6
    return 5


def write_option_pie_chart(
    ws,
    data_start_row: int,
    option: str,
    members: list[Student],
    anchor_cell: str,
    data_start_col: int,
    charts_in_row: int,
) -> int:
    option_name = BALANCE_OPTIONS[option]
    header_row = data_start_row
    ws.cell(row=header_row, column=data_start_col, value="类别")
    ws.cell(row=header_row, column=data_start_col + 1, value="人数")

    counts = Counter(balance_value(student, option) for student in members)
    if not counts:
        counts["无学生"] = 0
    current_row = header_row
    for label, count in sorted(counts.items(), key=lambda item: item[0]):
        current_row += 1
        ws.cell(row=current_row, column=data_start_col, value=label)
        ws.cell(row=current_row, column=data_start_col + 1, value=count)

    for col in (data_start_col, data_start_col + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    chart = PieChart()
    chart.title = f"{option_name}比例"
    chart.width, chart.height = pie_chart_size(len(counts), charts_in_row)
    chart.legend = None
    chart.visible_cells_only = False
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = True
    chart.dataLabels.showPercent = True
    chart.dataLabels.showLeaderLines = True
    chart.dataLabels.position = "bestFit"
    chart.dataLabels.separator = "\n"
    data = Reference(ws, min_col=data_start_col + 1, min_row=header_row, max_row=current_row)
    labels = Reference(ws, min_col=data_start_col, min_row=header_row + 1, max_row=current_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, anchor_cell)
    return current_row


def write_grouped_compact_roster(
    ws,
    start_row: int,
    by_class: dict[str, list[Student]],
    class_names: list[str],
    skipped_students: list[Student],
    balance_options: str,
) -> tuple[int, int, int]:
    ws.cell(row=start_row, column=1, value="班级名单（按班级分段）")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
    headers = compact_roster_headers()
    first_header_row = 0
    last_student_row = start_row
    current_row = start_row

    for class_name in class_names:
        current_row += 2
        ws.cell(row=current_row, column=1, value=class_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="1F4E78")
        ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="EAF3F8")
        current_row += 1
        if not first_header_row:
            first_header_row = current_row
        for col, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col, value=header)
        add_header_style(ws, current_row)

        members = by_class[class_name]
        class_header_row = current_row
        for idx, student in enumerate(members, start=1):
            current_row += 1
            for col, value in enumerate(compact_roster_values(student, idx, balance_options), start=1):
                ws.cell(row=current_row, column=col, value=value)
        if current_row == class_header_row:
            current_row += 1
            ws.cell(row=current_row, column=1, value=class_name)
            ws.cell(row=current_row, column=3, value="本班暂无学生")
        apply_table_borders(ws, class_header_row, current_row, 1, len(headers))
        last_student_row = current_row

        if balance_options:
            chart_row = current_row + 2
            helper_start_col = 60
            anchor_step = pie_chart_anchor_step(len(balance_options))
            for option_index, option in enumerate(balance_options):
                anchor_col = 1 + option_index * anchor_step
                data_col = helper_start_col + option_index * 3
                write_option_pie_chart(
                    ws=ws,
                    data_start_row=chart_row,
                    option=option,
                    members=members,
                    anchor_cell=f"{get_column_letter(anchor_col)}{chart_row}",
                    data_start_col=data_col,
                    charts_in_row=len(balance_options),
                )
            current_row = chart_row + 12

    if skipped_students:
        current_row += 2
        ws.cell(row=current_row, column=1, value="未参与分班")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="9C5700")
        ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
        current_row += 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col, value=header)
        add_header_style(ws, current_row)
        skipped_header_row = current_row
        for student in skipped_students:
            current_row += 1
            for col, value in enumerate(compact_roster_values(student, "", balance_options), start=1):
                ws.cell(row=current_row, column=col, value=value)
        apply_table_borders(ws, skipped_header_row, current_row, 1, len(headers))

    return current_row, first_header_row, last_student_row


def write_grouped_class_roster(
    ws,
    start_row: int,
    headers: list[str],
    by_class: dict[str, list[Student]],
    class_names: list[str],
    skipped_students: list[Student],
    balance_options: str,
) -> int:
    style_section_title(ws, start_row, "按班级名单", max_col=min(len(headers), 8))
    current_row = start_row
    for class_name in class_names:
        current_row += 2
        ws.cell(row=current_row, column=1, value=class_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="1F4E78")
        ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="EAF3F8")
        current_row += 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col, value=header)
        add_header_style(ws, current_row)
        for idx, student in enumerate(by_class[class_name], start=1):
            current_row += 1
            for col, value in enumerate(result_row_values(student, headers, idx, balance_options), start=1):
                ws.cell(row=current_row, column=col, value=value)

    if skipped_students:
        current_row += 2
        ws.cell(row=current_row, column=1, value="未参与分班")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="9C5700")
        ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
        current_row += 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col, value=header)
        add_header_style(ws, current_row)
        for student in skipped_students:
            current_row += 1
            for col, value in enumerate(result_row_values(student, headers, "", balance_options), start=1):
                ws.cell(row=current_row, column=col, value=value)
    return current_row


def write_result_workbook(
    source_path: Path,
    output_path: Path,
    original_headers: list[str],
    students: list[Student],
    class_names: list[str],
    seed: int,
    capacity: int | None,
    balance_options: str,
    issues: list[Issue],
    notices: list[str] | None = None,
    fallback_dir: Path | None = None,
) -> Path:
    wb = load_workbook_for_write(source_path)
    if RESULT_SHEET_NAME in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[RESULT_SHEET_NAME]
    result_ws = wb.create_sheet(RESULT_SHEET_NAME)

    result_headers = result_headers_for(original_headers)
    header_row = 1
    if notices:
        notice_text = "说明：" + "；".join(notices)
        result_ws.append([notice_text])
        result_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(result_headers))
        result_ws["A1"].font = Font(bold=True, color="9C5700")
        result_ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
        result_ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
        result_ws.row_dimensions[1].height = 36
        header_row = 2
    result_ws.append(result_headers)

    class_order = {name: idx for idx, name in enumerate(class_names)}
    participating_students = [student for student in students if student.participates]
    skipped_students = [student for student in students if not student.participates]
    sorted_participating = sorted(
        participating_students,
        key=lambda item: (
            class_order.get(item.assigned_class, 999),
            balance_layer(item, balance_options),
            item.student_id,
            item.name,
            item.row_number,
        ),
    )
    sorted_skipped = sorted(skipped_students, key=lambda item: (item.row_number, item.name))
    sorted_students = sorted_participating + sorted_skipped
    class_positions: Counter[str] = Counter()
    for student in sorted_students:
        class_index = ""
        if student.participates:
            class_positions[student.assigned_class] += 1
            class_index = class_positions[student.assigned_class]
        result_ws.append(result_row_values(student, result_headers, class_index, balance_options))
    add_header_style(result_ws, header_row)
    result_table_end_row = result_ws.max_row
    result_ws.freeze_panes = f"A{header_row + 1}"
    result_ws.auto_filter.ref = f"A{header_row}:{get_column_letter(result_ws.max_column)}{result_table_end_row}"

    by_class = {name: [student for student in sorted_participating if student.assigned_class == name] for name in class_names}
    current_row, compact_header_row, compact_roster_end_row = write_grouped_compact_roster(
        ws=result_ws,
        start_row=result_ws.max_row + 2,
        by_class=by_class,
        class_names=class_names,
        skipped_students=sorted_skipped,
        balance_options=balance_options,
    )

    stats_start_row = current_row + 2
    result_ws.cell(row=stats_start_row, column=1, value="班级统计")
    result_ws.cell(row=stats_start_row, column=1).font = Font(bold=True, size=12)
    stats_headers = ["班级", "人数", "男", "女", "上海市户口", "外省市户口"] + HUKOU_DETAILS + ["定向人数"]
    stats_header_row = stats_start_row + 1
    for col, header in enumerate(stats_headers, start=1):
        result_ws.cell(row=stats_header_row, column=col, value=header)
    totals = Counter()
    current_row = stats_header_row
    for class_name in class_names:
        members = by_class[class_name]
        row = [
            class_name,
            len(members),
            sum(1 for item in members if item.gender == "男"),
            sum(1 for item in members if item.gender == "女"),
            sum(1 for item in members if item.hukou_main == "上海市户口"),
            sum(1 for item in members if item.hukou_main == "外省市户口"),
        ]
        row.extend(sum(1 for item in members if item.hukou_detail == detail) for detail in HUKOU_DETAILS)
        row.append(sum(1 for item in members if item.directed))
        current_row += 1
        for col, value in enumerate(row, start=1):
            result_ws.cell(row=current_row, column=col, value=value)
        for header, value in zip(stats_headers[1:], row[1:]):
            totals[header] += value
    stats_last_class_row = current_row
    current_row += 1
    for col, value in enumerate(["合计"] + [totals[header] for header in stats_headers[1:]], start=1):
        result_ws.cell(row=current_row, column=col, value=value)
    for col in range(1, len(stats_headers) + 1):
        result_ws.cell(row=current_row, column=col).font = Font(bold=True)
        result_ws.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor="E2F0D9")

    info_start_row = current_row + 2
    result_ws.cell(row=info_start_row, column=1, value="分班说明")
    result_ws.cell(row=info_start_row, column=1).font = Font(bold=True, size=12)
    summary_rows = [
        ["分班依据", f"本次根据 {human_join(balance_option_names(balance_options))} 尽量平均分班。"],
        ["分班方法", "先安排指定班级的学生，再将其余学生按分班依据分组、随机打乱，并尽量平均放入各班。"],
        ["学生情况", f"读取 {len(students)} 行；参与分班 {len(participating_students)} 人；未参与分班 {len(skipped_students)} 人。"],
        ["班级情况", f"共 {len(class_names)} 个班：{'，'.join(class_names)}。"],
        ["每班上限", capacity if capacity is not None else "未设置"],
        ["随机种子", seed],
    ]
    if notices:
        summary_rows.append(["自动调整", "；".join(notices)])
    summary_header_row = info_start_row + 1
    for col, header in enumerate(["项目", "说明"], start=1):
        result_ws.cell(row=summary_header_row, column=col, value=header)
    current_row = summary_header_row
    for row in summary_rows:
        current_row += 1
        for col, value in enumerate(row, start=1):
            result_ws.cell(row=current_row, column=col, value=value)
            result_ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical="center")
    summary_end_row = current_row

    legacy_info_start_row = current_row + 2
    result_ws.cell(row=legacy_info_start_row, column=1, value="运行信息")
    result_ws.cell(row=legacy_info_start_row, column=1).font = Font(bold=True, size=12)
    info_rows = [
        ["生成时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["随机种子", seed],
        ["分班依据", balance_option_text(balance_options)],
        ["班级名称", "，".join(class_names)],
        ["每班人数上限", capacity if capacity is not None else "未设置"],
        ["学生总数", len(students)],
        ["参与分班人数", len(participating_students)],
        ["未参与分班人数", len(skipped_students)],
    ]
    legacy_info_header_row = legacy_info_start_row + 1
    for col, header in enumerate(["项目", "值"], start=1):
        result_ws.cell(row=legacy_info_header_row, column=col, value=header)
    current_row = legacy_info_header_row
    for row_offset, row in enumerate(info_rows):
        for col, value in enumerate(row, start=1):
            result_ws.cell(row=current_row + 1 + row_offset, column=col, value=value)
    legacy_info_end_row = current_row + len(info_rows)

    issue_start_row = legacy_info_end_row + 2
    result_ws.cell(row=issue_start_row, column=1, value="异常和跳过说明")
    result_ws.cell(row=issue_start_row, column=1).font = Font(bold=True, size=12)
    issue_header_row = issue_start_row + 1
    for col, value in enumerate(["操作", "原始行号", "学生姓名", "字段", "说明"], start=1):
        result_ws.cell(row=issue_header_row, column=col, value=value)
    names_by_row = student_name_by_row(students)
    for row_offset, issue in enumerate(issues or [Issue("提示", "-", "-", "未发现异常。")], start=1):
        result_ws.cell(row=issue_header_row + row_offset, column=1, value=issue.level)
        result_ws.cell(row=issue_header_row + row_offset, column=2, value=issue.row_number)
        result_ws.cell(row=issue_header_row + row_offset, column=3, value=issue_student_name(issue, names_by_row))
        result_ws.cell(row=issue_header_row + row_offset, column=4, value=issue.field)
        result_ws.cell(row=issue_header_row + row_offset, column=5, value=issue.message)
    issue_end_row = issue_header_row + len(issues or [Issue("提示", "-", "-", "未发现异常。")])

    add_header_style(result_ws, header_row)
    add_header_style(result_ws, stats_header_row)
    add_header_style(result_ws, summary_header_row)
    add_header_style(result_ws, legacy_info_header_row)
    add_header_style(result_ws, issue_header_row)
    apply_table_borders(result_ws, stats_header_row, stats_last_class_row + 1, 1, len(stats_headers))
    apply_table_borders(result_ws, summary_header_row, summary_end_row, 1, 2)
    apply_table_borders(result_ws, legacy_info_header_row, legacy_info_end_row, 1, 2)
    apply_table_borders(result_ws, issue_header_row, issue_end_row, 1, 5)
    autofit(result_ws, max_width=42)
    result_path = save_workbook_safely(wb, output_path, fallback_dir, allow_alternate=(source_path != output_path))
    wb.close()
    gc.collect()
    return result_path


def split_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    class_count: int | None = None,
    class_names: list[str] | None = None,
    seed: int | None = None,
    capacity: int | None = None,
    sheet_name: str | None = None,
    balance_options: str | None = None,
) -> SplitResult:
    input_path = Path(input_path)
    if not input_path.exists():
        raise ClassSplitError(f"找不到输入文件：{input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=False)
    try:
        config = read_config(wb)
    finally:
        wb.close()
        gc.collect()
    resolved_class_names = resolve_class_names(config, class_count, class_names)
    resolved_seed = seed if seed is not None else int_or_none(config.get("随机种子", ""))
    if resolved_seed is None:
        resolved_seed = int(dt.datetime.now().timestamp() * 1000) % 1_000_000_000
    resolved_capacity = capacity if capacity is not None else int_or_none(config.get("每班人数上限", ""))
    resolved_balance_options = parse_balance_options(balance_options or config.get("分班依据", ""))

    if output_path is None:
        output_path = input_path
    output_path = Path(output_path)
    fallback_dir = Path(__file__).resolve().parent

    students, original_headers, issues = read_students(input_path, sheet_name, resolved_class_names, resolved_balance_options)
    participating_students = [student for student in students if student.participates]
    resolved_class_names, notices = expand_classes_for_capacity(
        resolved_class_names,
        resolved_capacity,
        len(participating_students),
    )
    if students and not participating_students:
        issues.append(Issue("错误", "-", "学生名单", "没有可参与分班的有效学生；请检查姓名、性别和所选分班依据需要的字段。"))

    if has_errors(issues):
        report_path = write_validation_report(output_path, original_headers, students, issues, fallback_dir)
        raise ClassSplitError(f"名单校验失败，已生成异常报告：{report_path}", report_path)

    assign_students(participating_students, resolved_class_names, resolved_seed, resolved_capacity, resolved_balance_options, issues)
    if has_errors(issues):
        report_path = write_validation_report(output_path, original_headers, students, issues, fallback_dir)
        raise ClassSplitError(f"分班失败，已生成异常报告：{report_path}", report_path)

    final_output_path = write_result_workbook(
        source_path=input_path,
        output_path=output_path,
        original_headers=original_headers,
        students=students,
        class_names=resolved_class_names,
        seed=resolved_seed,
        capacity=resolved_capacity,
        balance_options=resolved_balance_options,
        issues=issues,
        notices=notices,
        fallback_dir=fallback_dir,
    )
    return SplitResult(final_output_path, len(students), resolved_class_names, resolved_seed, resolved_balance_options, issues)


def create_template(path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "学生名单"
    ws.append(TEMPLATE_COLUMNS)
    add_header_style(ws)
    ws.freeze_panes = "A2"

    def column_range(header: str, start_row: int = 2, end_row: int = 1000) -> str:
        col_idx = TEMPLATE_COLUMNS.index(header) + 1
        letter = get_column_letter(col_idx)
        return f"{letter}{start_row}:{letter}{end_row}"

    gender_dv = DataValidation(type="list", formula1='"男,女"', allow_blank=False)
    main_dv = DataValidation(type="list", formula1='"上海市户口,外省市户口"', allow_blank=False)
    detail_dv = DataValidation(type="list", formula1='"本市人户一致,本市人户分离,外省市满积分,外省市不满积分"', allow_blank=False)
    yes_no_dv = DataValidation(type="list", formula1='"否,是"', allow_blank=True)
    for dv in [gender_dv, main_dv, detail_dv, yes_no_dv]:
        ws.add_data_validation(dv)
    gender_dv.add(column_range("性别"))
    main_dv.add(column_range("户籍大类"))
    detail_dv.add(column_range("户籍细类"))
    yes_no_dv.add(column_range("是否定向"))

    ws.cell(row=1, column=TEMPLATE_COLUMNS.index("总人数序号") + 1).comment = Comment("原始报名导出字段可以直接粘贴进来。", "Codex")
    ws.cell(row=1, column=TEMPLATE_COLUMNS.index("儿童人员类型") + 1).comment = Comment("选择 D 分班时会使用此字段。", "Codex")
    ws.cell(row=1, column=TEMPLATE_COLUMNS.index("户籍细类") + 1).comment = Comment("可人工覆盖推断结果；留空时工具会根据户籍、居住、人户分离和积分关键词推断。", "Codex")
    ws.cell(row=1, column=TEMPLATE_COLUMNS.index("指定班级") + 1).comment = Comment("需要定向安排时填写班级名称，例如：2班。", "Codex")
    autofit(ws, max_width=26)

    config = wb.create_sheet("配置")
    config.append(["参数", "值", "说明"])
    config.append(["班级数量", 4, "不填写班级名称时，自动生成 1班、2班……"])
    config.append(["班级名称", "1班，2班，3班，4班", "建议用中文逗号“，”分隔；名称个数要和班级数量一致。"])
    config.append(["分班依据", DEFAULT_BALANCE_OPTIONS, "输入 A/B/C/D 组合：A性别，B户籍大类，C户籍细类，D儿童人员类型。"])
    config.append(["随机种子", "", "留空则每次随机；填写数字则结果可复现。"])
    config.append(["每班人数上限", "", "留空表示不设置硬上限。"])
    add_header_style(config)
    config.column_dimensions["A"].width = 16
    config.column_dimensions["B"].width = 30
    config.column_dimensions["C"].width = 46

    notes = wb.create_sheet("说明")
    notes.append(["步骤", "说明"])
    notes.append(["1", "在“学生名单”中填写学生信息。"])
    notes.append(["2", "如需定向安排，在“是否定向”填“是”，并在“指定班级”填写班级名称。"])
    notes.append(["3", "双击“启动分班工具.bat”，在“根据哪些项分班”输入 A、AB、ABC 或 ABCD。"])
    notes.append(["4", "工具会在原 Excel 中新增或更新“分班结果”页面，不另存新文件。"])
    notes.append(["每班上限", "如果每班上限乘以班级数量容纳不了学生，会自动新增班级，并在“分班结果”第一行说明。"])
    notes.append(["分班依据", "A=性别，B=户籍大类，C=户籍细类，D=儿童人员类型。"])
    notes.append(["户籍细类", "本市人户一致、本市人户分离、外省市满积分、外省市不满积分。"])
    add_header_style(notes)
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 72

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def run_self_test() -> None:
    with tempfile.TemporaryDirectory(prefix="class_splitter_") as tmp:
        tmp_path = Path(tmp)
        input_path = tmp_path / "selftest_input.xlsx"
        create_template(input_path)
        wb = load_workbook(input_path)
        ws = wb["学生名单"]
        details = HUKOU_DETAILS * 8
        genders = (["男", "女"] * 16)[:32]
        headers = [cell.value for cell in ws[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers)}
        for idx in range(32):
            detail = details[idx]
            main = HUKOU_MAIN_BY_DETAIL[detail]
            directed = "是" if idx in {0, 7} else "否"
            target = "1班" if idx == 0 else ("4班" if idx == 7 else "")
            row_num = idx + 2
            values = {
                "总人数序号": f"{idx + 1:03d}",
                "姓名": f"学生{idx + 1:02d}",
                "性别": genders[idx],
                "儿童人员类型": "外省市满积分" if detail == "外省市满积分" else ("外省市不满积分" if detail == "外省市不满积分" else "本市户籍"),
                "户籍地址(省)": "上海市" if main == "上海市户口" else "江苏省",
                "户籍地址(区)": "浦东新区" if detail == "本市人户一致" else "徐汇区",
                "居住地址(区)": "浦东新区",
                "儿童人户分离登记情况": "人户分离" if detail == "本市人户分离" else ("人户一致" if detail == "本市人户一致" else ""),
                "是否定向": directed,
                "指定班级": target,
                "户籍大类": main,
                "户籍细类": detail,
            }
            for header, value in values.items():
                ws.cell(row=row_num, column=header_index[header], value=value)
        bad_rows = [
            {"总人数序号": "900", "姓名": "缺性别学生", "户籍地址(省)": "上海市", "户籍细类": "本市人户一致"},
            {"总人数序号": "901", "性别": "男性", "户籍地址(省)": "上海市", "户籍细类": "本市人户一致"},
        ]
        for offset, values in enumerate(bad_rows, start=34):
            for header, value in values.items():
                ws.cell(row=offset, column=header_index[header], value=value)
        original_sheet_checks = {
            "A2": ws["A2"].value,
            "C2": ws["C2"].value,
            "I2": ws["I2"].value,
            "CD2": ws["CD2"].value,
        }
        wb.save(input_path)

        result = split_workbook(input_path, class_count=4, seed=20260627, capacity=7, balance_options="ABC")
        result_wb = load_workbook(result.output_path, data_only=True)
        original_ws_after = result_wb["学生名单"]
        for address, expected_value in original_sheet_checks.items():
            if original_ws_after[address].value != expected_value:
                raise AssertionError("自检失败：原学生名单页被修改。")
        result_ws = result_wb["分班结果"]
        if "5班" not in result.class_names:
            raise AssertionError("自检失败：容量不足时未自动新增班级。")
        if not clean(result_ws["A1"].value).startswith("说明：因每班上限"):
            raise AssertionError("自检失败：未在第一行写入自动新增班级说明。")
        expected_chart_count = len(result.class_names) * len(result.balance_options)
        if len(result_ws._charts) < expected_chart_count:
            raise AssertionError("自检失败：未按每班和所选依据生成饼图。")
        first_chart = result_ws._charts[0]
        if (
            not first_chart.dataLabels
            or not first_chart.dataLabels.showCatName
            or not first_chart.dataLabels.showVal
            or not first_chart.dataLabels.showPercent
        ):
            raise AssertionError("自检失败：饼图未将类别、人数和比例显示在图上。")
        if first_chart.visible_cells_only:
            raise AssertionError("自检失败：饼图只读取可见单元格，隐藏辅助数据时会显示为空白。")
        if first_chart.legend is not None:
            raise AssertionError("自检失败：饼图未关闭图例，容易挤占标签空间。")
        if first_chart.dataLabels.separator != "\n":
            raise AssertionError("自检失败：饼图标签未换行显示，可能不够清晰。")
        test_width, test_height = pie_chart_size(category_count=8, charts_in_row=4)
        if test_width > 5.8 or test_height > 5.4 or test_width < 4.6 or test_height < 4.3:
            raise AssertionError("自检失败：饼图自动尺寸超出限制。")
        if pie_chart_anchor_step(4) != 5 or pie_chart_anchor_step(2) != 6:
            raise AssertionError("自检失败：饼图横向间距规则不正确。")
        if not result_ws.column_dimensions[get_column_letter(60)].hidden:
            raise AssertionError("自检失败：饼图辅助数据列未隐藏。")
        assigned_rows = []
        for row in result_ws.iter_rows(min_row=3, values_only=True):
            if not any(value is not None for value in row):
                break
            assigned_rows.append(row)
        if len(assigned_rows) != 34:
            raise AssertionError("自检失败：结果人数不正确。")
        for sheet_name in ["班级学生名单", "1班名单", "2班名单", "3班名单", "4班名单", "未参与分班", "班级统计", "运行信息", "异常校验"]:
            if sheet_name in result_wb.sheetnames:
                raise AssertionError(f"自检失败：不应再生成工作表 {sheet_name}。")
        headers = [cell.value for cell in result_ws[2]]
        class_col = headers.index("最终班级")
        counts = Counter(row[class_col] for row in assigned_rows if row[class_col] != "未参与分班")
        if max(counts.values()) - min(counts.values()) > 1:
            raise AssertionError("自检失败：班级人数差异超过1。")
        if max(counts.values()) > 7:
            raise AssertionError("自检失败：班级人数超过上限。")
        skipped_count = sum(1 for row in assigned_rows if row[class_col] == "未参与分班")
        if skipped_count != 2:
            raise AssertionError("自检失败：未参与分班人数不正确。")
        found_summary = any(
            clean(cell.value) == "分班说明"
            for row in result_ws.iter_rows(values_only=False)
            for cell in row
        )
        if not found_summary:
            raise AssertionError("自检失败：未生成分班说明。")
        found_grouped_roster = any(
            clean(cell.value) == "班级名单（按班级分段）"
            for row in result_ws.iter_rows(values_only=False)
            for cell in row
        )
        if not found_grouped_roster:
            raise AssertionError("自检失败：未生成按班级分段的班级名单。")
        print(f"自检通过：{result.output_path}")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按性别和户籍信息均衡随机分班。")
    parser.add_argument("input", nargs="?", help="输入 Excel 名单路径。")
    parser.add_argument("-o", "--output", help="可选：另存到指定 Excel 路径。默认写回原文件的“分班结果”页面。")
    parser.add_argument("--class-count", type=int, help="班级数量，例如 4。")
    parser.add_argument("--classes", help="班级名称，例如：1班,2班,3班,4班。")
    parser.add_argument("--seed", type=int, help="随机种子。填写后结果可复现。")
    parser.add_argument("--capacity", type=int, help="每班人数上限。")
    parser.add_argument("--balance", help="分班依据，输入 A/B/C/D 组合。A=性别，B=户籍大类，C=户籍细类，D=儿童人员类型。默认 ABC。")
    parser.add_argument("--sheet", help="学生名单所在工作表名称。默认优先使用“学生名单”。")
    parser.add_argument("--create-template", help="生成标准名单模板。")
    parser.add_argument("--self-test", action="store_true", help="运行内置自检。")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])

    if args.self_test:
        run_self_test()
        return 0

    if args.create_template:
        path = create_template(args.create_template)
        print(f"已生成模板：{path}")
        return 0

    if not args.input:
        print("请提供输入 Excel 文件，或使用 --create-template 生成模板。", file=sys.stderr)
        return 2

    class_names = split_csv_like(args.classes) if args.classes else None
    try:
        result = split_workbook(
            input_path=args.input,
            output_path=args.output,
            class_count=args.class_count,
            class_names=class_names,
            seed=args.seed,
            capacity=args.capacity,
            sheet_name=args.sheet,
            balance_options=args.balance,
        )
    except ClassSplitError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    print(f"分班完成：{result.output_path}")
    print(f"学生人数：{result.student_count}")
    print(f"班级：{'，'.join(result.class_names)}")
    print(f"随机种子：{result.seed}")
    print(f"分班依据：{balance_option_text(result.balance_options)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
